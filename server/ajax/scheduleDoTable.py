import os
import datetime
import pathlib
import csv

# ------------------------------------------------------------------

from sqlalchemy import exc

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from database.tables import User, InTag, Grid, Spindle, SpindleRunIn, RunInData, Session
from flask import Blueprint, jsonify, request, current_app

from apscheduler.schedulers.background import BackgroundScheduler

from dotenv import load_dotenv

scheduleDoTable = Blueprint('scheduleDoTable', __name__)

file_ok = False

# ------------------------------------------------------------------

def do_read_all_excel_files():
  global file_ok

  print("do_read_all_excel_files()....")

  # 指定目录路径
  #_base_dir = 'd:\\釸達輔導案\開發文件\跑合機匯出EXCEL'
  #---
  #_base_dir = current_app.config['baseDir']

  load_dotenv()
  _base_dir = os.getenv("baseDir2")
  #---

  # 讀取指定目錄下的所有指定檔案名稱
  files = [f for f in os.listdir(_base_dir) if os.path.isfile(os.path.join(_base_dir, f)) and f.startswith('Report_') and f.endswith('.xlsx')]

  _customer_row = current_app.config['customer_row']
  _spindle_cat_row = current_app.config['spindle_cat_row']
  _id_row = current_app.config['id_row']
  _id_column = current_app.config['id_column']
  _emp_id_row = current_app.config['emp_id_row']
  _date_row = current_app.config['date_row']
  _date_column = current_app.config['date_column']
  _runin_row = current_app.config['runin_row']

  s = Session()

  for _file_name in files:
    print(_file_name)
    existing_excel = s.query(SpindleRunIn).filter_by(spindleRunIn_excel_file = _file_name).first()
    _path = _base_dir + '\\' + _file_name
    workbook = openpyxl.load_workbook(filename = _path, read_only = True)

    if existing_excel or ('Sheet1' not in workbook.sheetnames):
      return_message1 = '錯誤! excel檔名重複或檔案內沒有Sheet1...'
      print(return_message1)
      continue
    #continue for loop
    print('Sheet1 exists')
    sheet = workbook.active   #取得第1個工作表
    _spindleRunIn_customer = sheet.cell(row = _customer_row, column = 1).value    #客戶資料
    _cat = str(sheet.cell(row = _spindle_cat_row, column = 1).value).strip()                   #主軸資料(spindle_cat)
    _spindleRunIn_work_id = sheet.cell(row = _id_row, column = _id_column).value  #工單
    _empID = str(sheet.cell(row = _emp_id_row, column = 1).value).strip().zfill(4)                 #員工資料
    _spindleRunIn_date = str(sheet.cell(row = _date_row, column = _date_column).value.date()) #測試日期
    print('read Sheet1 upper part ok...', _cat, _empID)
    _spindle = s.query(Spindle).filter_by(spindle_cat = _cat).first()
    _user = s.query(User).filter_by(emp_id = _empID).first()

    if (not _spindle or not _user):
      return_message2 = '錯誤! 在excel檔案內, 系統沒有主軸或員工編號資料...'
      print(return_message2)
      continue
    #continue for loop
    print('write data into SpindleRunIn table...')
    new_spindle_runin = SpindleRunIn(
      spindleRunIn_excel_file = _file_name,
      spindleRunIn_customer = _spindleRunIn_customer,
      spindleRunIn_work_id = _spindleRunIn_work_id,
      spindleRunIn_spindle_id = _spindle.id,
      spindleRunIn_employer = _user.id,
      spindleRunIn_date = _spindleRunIn_date,
    )
    s.add(new_spindle_runin)

    s.flush()
    spindle_runin_id = new_spindle_runin.id
    print("spindle_runin_id: ", spindle_runin_id)
    #continue for loop
    print("read data from excel file...")
    _results = []
    for i in range(_runin_row, sheet.max_row+1):
      if sheet.cell(row = i, column = 1).value is None:
        break

      _obj = {
        'spindleRunIn_period': sheet.cell(row = i, column = 1).value,         #時間
        'spindleRunIn_speed_level': sheet.cell(row = i, column = 2).value,    #段速
        'spindleRunIn_speed': sheet.cell(row = i, column = 3).value,          #轉速
        'spindleRunIn_stator_temp': sheet.cell(row = i, column = 4).value,    #定子溫度
        'spindleRunIn_inner_frontBearing_temp': sheet.cell(row = i, column = 5).value,   #內部前軸承溫度
        'spindleRunIn_inner_backBearing_temp': sheet.cell(row = i, column = 6).value,   #內部後軸承溫度
        'spindleRunIn_outer_frontBearing_temp': sheet.cell(row = i, column = 7).value,  #外部前軸承溫度
        'spindleRunIn_outer_backBearing_temp': sheet.cell(row = i, column = 8).value,   #外部後軸承溫度
        'spindleRunIn_room_temp': sheet.cell(row = i, column = 9).value,          #室溫
        'spindleRunIn_coolWater_temp': sheet.cell(row = i, column = 10).value,    #冷卻機水溫
        'spindleRunIn_Rphase_current': sheet.cell(row = i, column = 11).value,    #R相電流
        'spindleRunIn_Sphase_current': sheet.cell(row = i, column = 12).value,    #S相電流
        'spindleRunIn_Tphase_current': sheet.cell(row = i, column = 13).value,    #T相電流
        'spindleRunIn_cool_pipeline_flow': sheet.cell(row = i, column = 14).value,              #冷卻機管路流量
        'spindleRunIn_cool_pipeline_pressure': sheet.cell(row = i, column = 15).value,          #冷卻機管路壓力
        'spindleRunIn_frontBearing_vibration_speed1': sheet.cell(row = i, column = 16).value,   #前軸承震動速度-1
        'spindleRunIn_frontBearing_vibration_acc1': sheet.cell(row = i, column = 17).value,     #前軸承震動加速度-1
        'spindleRunIn_frontBearing_vibration_disp1': sheet.cell(row = i, column = 18).value,    #前軸承震動位移-1
        'spindleRunIn_frontBearing_vibration_speed2': sheet.cell(row = i, column = 19).value,   #前軸承震動速度-2
        'spindleRunIn_frontBearing_vibration_acc2': sheet.cell(row = i, column = 20).value,     #前軸承震動加速度-2
        'spindleRunIn_frontBearing_vibration_disp2': sheet.cell(row = i, column = 21).value,    #前軸承震動位移-2
        'spindleRunIn_backBearing_vibration_speed1': sheet.cell(row = i, column = 22).value,    #後軸承震動速度-1
        'spindleRunIn_backBearing_vibration_acc1': sheet.cell(row = i, column = 23).value,      #後軸承震動加速度-1
        'spindleRunIn_backBearing_vibration_disp1': sheet.cell(row = i, column = 24).value,     #後軸承震動位移-1
        'spindleRunIn_backBearing_vibration_speed2': sheet.cell(row = i, column = 25).value,    #後軸承震動速度-2
        'spindleRunIn_backBearing_vibration_acc2': sheet.cell(row = i, column = 26).value,      #後軸承震動加速度-2
        'spindleRunIn_backBearing_vibration_disp2': sheet.cell(row = i, column = 27).value,     #後軸承震動位移-2
      }
      _results.append(_obj)
    #continue for loop
    print('write data into RunInData table...')
    runin_data_total_size = len(_results)
    _objects = []
    for x in range(runin_data_total_size):
      u = RunInData(
      spindleRunIn_id = spindle_runin_id,
      spindleRunIn_period = _results[x]['spindleRunIn_period'],
      spindleRunIn_speed_level = _results[x]['spindleRunIn_speed_level'],
      spindleRunIn_speed = _results[x]['spindleRunIn_speed'],
      spindleRunIn_stator_temp = _results[x]['spindleRunIn_stator_temp'],
      spindleRunIn_inner_frontBearing_temp = _results[x]['spindleRunIn_inner_frontBearing_temp'],
      spindleRunIn_inner_backBearing_temp = _results[x]['spindleRunIn_inner_backBearing_temp'],
      spindleRunIn_outer_frontBearing_temp = _results[x]['spindleRunIn_outer_frontBearing_temp'],
      spindleRunIn_outer_backBearing_temp = _results[x]['spindleRunIn_outer_backBearing_temp'],
      spindleRunIn_room_temp = _results[x]['spindleRunIn_room_temp'],
      spindleRunIn_coolWater_temp = _results[x]['spindleRunIn_coolWater_temp'],
      spindleRunIn_Rphase_current = _results[x]['spindleRunIn_Rphase_current'],
      spindleRunIn_Sphase_current = _results[x]['spindleRunIn_Sphase_current'],
      spindleRunIn_Tphase_current = _results[x]['spindleRunIn_Tphase_current'],
      spindleRunIn_cool_pipeline_flow = _results[x]['spindleRunIn_cool_pipeline_flow'],
      spindleRunIn_cool_pipeline_pressure = _results[x]['spindleRunIn_cool_pipeline_pressure'],
      spindleRunIn_frontBearing_vibration_speed1 = _results[x]['spindleRunIn_frontBearing_vibration_speed1'],
      spindleRunIn_frontBearing_vibration_acc1 = _results[x]['spindleRunIn_frontBearing_vibration_acc1'],
      spindleRunIn_frontBearing_vibration_disp1 = _results[x]['spindleRunIn_frontBearing_vibration_disp1'],
      spindleRunIn_frontBearing_vibration_speed2 = _results[x]['spindleRunIn_frontBearing_vibration_speed2'],
      spindleRunIn_frontBearing_vibration_acc2 = _results[x]['spindleRunIn_frontBearing_vibration_acc2'],
      spindleRunIn_frontBearing_vibration_disp2 = _results[x]['spindleRunIn_frontBearing_vibration_disp2'],
      spindleRunIn_backBearing_vibration_speed1 = _results[x]['spindleRunIn_backBearing_vibration_speed1'],
      spindleRunIn_backBearing_vibration_acc1 = _results[x]['spindleRunIn_backBearing_vibration_acc1'],
      spindleRunIn_backBearing_vibration_disp1 = _results[x]['spindleRunIn_backBearing_vibration_disp1'],
      spindleRunIn_backBearing_vibration_speed2 = _results[x]['spindleRunIn_backBearing_vibration_speed2'],
      spindleRunIn_backBearing_vibration_acc2 = _results[x]['spindleRunIn_backBearing_vibration_acc2'],
      spindleRunIn_backBearing_vibration_disp2 = _results[x]['spindleRunIn_backBearing_vibration_disp2'],
      )
      _objects.append(u)

    s.bulk_save_objects(_objects)

    try:
        s.commit()
    except pymysql.err.IntegrityError as e:
        s.rollback()
    except exc.IntegrityError as e:
        s.rollback()
    except Exception as e:
        s.rollback()
    #continue for loop
    print("spindle_runin and runin_data combine...")
    spindle_runin_record = s.query(SpindleRunIn).filter_by(id = spindle_runin_id).first()
    runin_data_records = s.query(RunInData).filter_by(spindleRunIn_id = spindle_runin_id).all()

    for array in runin_data_records:
      spindle_runin_record._runin_data.append(array)

    try:
        s.commit()
    except pymysql.err.IntegrityError as e:
        s.rollback()
    except exc.IntegrityError as e:
        s.rollback()
    except Exception as e:
        s.rollback()
    #end for loop

  s.close()
  #end do_read_all_excel_files()
  file_ok = True

'''
def my_job():
  print("hello, Scheduled job is running...")

  do_read_all_excel_files()
  #end my_job()

# 注册定时任务到调度器
scheduler = BackgroundScheduler()
scheduler.add_job(my_job, 'cron', hour=14, minute=40)

# 开始执行调度器中注册的任务
scheduler.start()
'''