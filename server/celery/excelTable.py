import os
import datetime
import pathlib
import csv

# ------------------------------------------------------------------

from sqlalchemy import exc


import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from database.tables import User, InTag, Grid, Spindle, SpindleRunIn, RunInData, Session
from flask import Blueprint, jsonify, request, current_app

from dotenv import load_dotenv

# ------------------------------------------------------------------

excelTable = Blueprint('excelTable', __name__)

def pad_zeros(str):
    #length = len(str)
    #if length <= 1:
    return str.zfill(4)
    #else:
    #    return str.zfill(length + 3)


@excelTable.route("/readAllExcelFiles", methods=['GET'])
def read_all_excel_files():
  print("readAllExcelFiles....")

  return_value = True
  return_message1 = ''
  return_message2 = ''

  _repeat_excel_files = []
  _no_spindle_or_user_data = []

  # 指定目錄path
  #_base_dir = 'd:\\釸達輔導案\開發文件\跑合機匯出EXCEL'
  #---
  #_base_dir = current_app.config['baseDir']

  #load_dotenv()
  #dotenv_path = "d:\\theta-asrs\\asrs\\.env"     #2024-04-26 add, mark
  #load_dotenv(dotenv_path)                             #2024-04-26 add, mark
  #_base_dir = os.getenv("baseDir")                 # 2024-04-26 mark
  _base_dir = current_app.config['baseDir']         # 2024-04-26 add

  print("read excel files, dir: ", _base_dir)       # 2024-04-26 add
  #---

  # 讀取指定目錄下的所有檔案和目錄名稱
  #files_and_folders = os.listdir(_base_dir)
  # 讀取指定目錄下的所有檔案名稱
  #files = [f for f in os.listdir(_base_dir) if os.path.isfile(os.path.join(_base_dir, f))]
  # 讀取指定目錄下的所有指定檔案名稱
  files = [f for f in os.listdir(_base_dir) if os.path.isfile(os.path.join(_base_dir, f)) and f.startswith('Report_') and f.endswith('.xlsx')]
  #print(files)
  _customer_row = current_app.config['customer_row']
  _spindle_cat_row = current_app.config['spindle_cat_row']
  _id_row = current_app.config['id_row']
  _id_column = current_app.config['id_column']
  _emp_id_row = current_app.config['emp_id_row']
  _date_row = current_app.config['date_row']
  _date_column = current_app.config['date_column']
  _runin_row = current_app.config['runin_row']

  s = Session()
  #for item in files_and_folders:
  for _file_name in files:
    print(_file_name)
    existing_excel = s.query(SpindleRunIn).filter_by(spindleRunIn_excel_file = _file_name).first()
    _path = _base_dir + '\\' + _file_name
    workbook = openpyxl.load_workbook(filename = _path, read_only = True)

    if existing_excel or ('Sheet1' not in workbook.sheetnames):
      #_repeat_excel_files.append(_file_name)
      #return_value = False
      return_message1 = '錯誤! excel檔名重複或檔案內沒有Sheet1...'
      print(return_message1)
      continue
    #continue for loop
    print('Sheet1 exists')
    sheet = workbook.active   #取得第1個工作表
    _spindleRunIn_customer = sheet.cell(row = _customer_row, column = 1).value    #客戶資料
    _cat = str(sheet.cell(row = _spindle_cat_row, column = 1).value).strip()                   #主軸資料(spindle_cat)
    _spindleRunIn_work_id = sheet.cell(row = _id_row, column = _id_column).value  #工單
    _empID = pad_zeros(str(sheet.cell(row = _emp_id_row, column = 1).value).strip())                 #員工資料
    _spindleRunIn_date = str(sheet.cell(row = _date_row, column = _date_column).value).strip() #測試日期
    print('read Sheet1 upper part ok...', _cat, _empID)
    _spindle = s.query(Spindle).filter_by(spindle_cat = _cat).first()
    _user = s.query(User).filter_by(emp_id = _empID).first()

    if (not _spindle or not _user):
      temp_data = _file_name + ': ' + _cat + ', ' + _empID
      _no_spindle_or_user_data.append(temp_data)
      return_value = False
      return_message2 = '錯誤! 在excel檔案內, 系統沒有主軸或員工編號資料...'
      print(return_message2)
      continue
    '''
    if (not _spindle):
      temp_data = _file_name + ': ' + _cat + ', ' + _empID
      _no_spindle_or_user_data.append(temp_data)
      return_value = False
      return_message2 = '錯誤! 在excel檔案內, 系統沒有主軸資料...'
      print(return_message2)
      continue
    if (not _user):
      temp_data = _file_name + ': ' + _cat + ', ' + _empID
      _no_spindle_or_user_data.append(temp_data)
      return_value = False
      return_message2 = '錯誤! 在excel檔案內, 系統沒有員工編號資料...'
      print(return_message2)
      continue
    '''
    #continue for loop
    print('write data into SpindleRunIn table...')
    new_spindle_runin = SpindleRunIn(
      spindleRunIn_excel_file = _file_name,
      spindleRunIn_customer = _spindleRunIn_customer,
      spindleRunIn_work_id = _spindleRunIn_work_id,
      spindleRunIn_spindle_id = _spindle.id,
      spindleRunIn_employer = _user.id,
      spindleRunIn_date = _spindleRunIn_date,
    )
    s.add(new_spindle_runin)

    s.flush()
    spindle_runin_id = new_spindle_runin.id
    print("spindle_runin_id: ", spindle_runin_id)
    #continue for loop
    print("read data from excel file...")
    _results = []
    for i in range(_runin_row, sheet.max_row+1):
      if sheet.cell(row = i, column = 1).value is None:
        break

      _obj = {
        'spindleRunIn_period': sheet.cell(row = i, column = 1).value,         #時間
        'spindleRunIn_speed_level': sheet.cell(row = i, column = 2).value,    #段速
        'spindleRunIn_speed': sheet.cell(row = i, column = 3).value,          #轉速
        'spindleRunIn_stator_temp': sheet.cell(row = i, column = 4).value,    #定子溫度
        'spindleRunIn_inner_frontBearing_temp': sheet.cell(row = i, column = 5).value,   #內部前軸承溫度
        'spindleRunIn_inner_backBearing_temp': sheet.cell(row = i, column = 6).value,   #內部後軸承溫度
        'spindleRunIn_outer_frontBearing_temp': sheet.cell(row = i, column = 7).value,  #外部前軸承溫度
        'spindleRunIn_outer_backBearing_temp': sheet.cell(row = i, column = 8).value,   #外部後軸承溫度
        'spindleRunIn_room_temp': sheet.cell(row = i, column = 9).value,          #室溫
        'spindleRunIn_coolWater_temp': sheet.cell(row = i, column = 10).value,    #冷卻機水溫
        'spindleRunIn_Rphase_current': sheet.cell(row = i, column = 11).value,    #R相電流
        'spindleRunIn_Sphase_current': sheet.cell(row = i, column = 12).value,    #S相電流
        'spindleRunIn_Tphase_current': sheet.cell(row = i, column = 13).value,    #T相電流
        'spindleRunIn_cool_pipeline_flow': sheet.cell(row = i, column = 14).value,              #冷卻機管路流量
        'spindleRunIn_cool_pipeline_pressure': sheet.cell(row = i, column = 15).value,          #冷卻機管路壓力
        'spindleRunIn_frontBearing_vibration_speed1': sheet.cell(row = i, column = 16).value,   #前軸承震動速度-1
        'spindleRunIn_frontBearing_vibration_acc1': sheet.cell(row = i, column = 17).value,     #前軸承震動加速度-1
        'spindleRunIn_frontBearing_vibration_disp1': sheet.cell(row = i, column = 18).value,    #前軸承震動位移-1
        'spindleRunIn_frontBearing_vibration_speed2': sheet.cell(row = i, column = 19).value,   #前軸承震動速度-2
        'spindleRunIn_frontBearing_vibration_acc2': sheet.cell(row = i, column = 20).value,     #前軸承震動加速度-2
        'spindleRunIn_frontBearing_vibration_disp2': sheet.cell(row = i, column = 21).value,    #前軸承震動位移-2
        'spindleRunIn_backBearing_vibration_speed1': sheet.cell(row = i, column = 22).value,    #後軸承震動速度-1
        'spindleRunIn_backBearing_vibration_acc1': sheet.cell(row = i, column = 23).value,      #後軸承震動加速度-1
        'spindleRunIn_backBearing_vibration_disp1': sheet.cell(row = i, column = 24).value,     #後軸承震動位移-1
        'spindleRunIn_backBearing_vibration_speed2': sheet.cell(row = i, column = 25).value,    #後軸承震動速度-2
        'spindleRunIn_backBearing_vibration_acc2': sheet.cell(row = i, column = 26).value,      #後軸承震動加速度-2
        'spindleRunIn_backBearing_vibration_disp2': sheet.cell(row = i, column = 27).value,     #後軸承震動位移-2
      }
      _results.append(_obj)
    #continue for loop
    print('write data into RunInData table...')
    runin_data_total_size = len(_results)
    _objects = []
    for x in range(runin_data_total_size):
      u = RunInData(
      spindleRunIn_id = spindle_runin_id,
      spindleRunIn_period = _results[x]['spindleRunIn_period'],
      spindleRunIn_speed_level = _results[x]['spindleRunIn_speed_level'],
      spindleRunIn_speed = _results[x]['spindleRunIn_speed'],
      spindleRunIn_stator_temp = _results[x]['spindleRunIn_stator_temp'],
      spindleRunIn_inner_frontBearing_temp = _results[x]['spindleRunIn_inner_frontBearing_temp'],
      spindleRunIn_inner_backBearing_temp = _results[x]['spindleRunIn_inner_backBearing_temp'],
      spindleRunIn_outer_frontBearing_temp = _results[x]['spindleRunIn_outer_frontBearing_temp'],
      spindleRunIn_outer_backBearing_temp = _results[x]['spindleRunIn_outer_backBearing_temp'],
      spindleRunIn_room_temp = _results[x]['spindleRunIn_room_temp'],
      spindleRunIn_coolWater_temp = _results[x]['spindleRunIn_coolWater_temp'],
      spindleRunIn_Rphase_current = _results[x]['spindleRunIn_Rphase_current'],
      spindleRunIn_Sphase_current = _results[x]['spindleRunIn_Sphase_current'],
      spindleRunIn_Tphase_current = _results[x]['spindleRunIn_Tphase_current'],
      spindleRunIn_cool_pipeline_flow = _results[x]['spindleRunIn_cool_pipeline_flow'],
      spindleRunIn_cool_pipeline_pressure = _results[x]['spindleRunIn_cool_pipeline_pressure'],
      spindleRunIn_frontBearing_vibration_speed1 = _results[x]['spindleRunIn_frontBearing_vibration_speed1'],
      spindleRunIn_frontBearing_vibration_acc1 = _results[x]['spindleRunIn_frontBearing_vibration_acc1'],
      spindleRunIn_frontBearing_vibration_disp1 = _results[x]['spindleRunIn_frontBearing_vibration_disp1'],
      spindleRunIn_frontBearing_vibration_speed2 = _results[x]['spindleRunIn_frontBearing_vibration_speed2'],
      spindleRunIn_frontBearing_vibration_acc2 = _results[x]['spindleRunIn_frontBearing_vibration_acc2'],
      spindleRunIn_frontBearing_vibration_disp2 = _results[x]['spindleRunIn_frontBearing_vibration_disp2'],
      spindleRunIn_backBearing_vibration_speed1 = _results[x]['spindleRunIn_backBearing_vibration_speed1'],
      spindleRunIn_backBearing_vibration_acc1 = _results[x]['spindleRunIn_backBearing_vibration_acc1'],
      spindleRunIn_backBearing_vibration_disp1 = _results[x]['spindleRunIn_backBearing_vibration_disp1'],
      spindleRunIn_backBearing_vibration_speed2 = _results[x]['spindleRunIn_backBearing_vibration_speed2'],
      spindleRunIn_backBearing_vibration_acc2 = _results[x]['spindleRunIn_backBearing_vibration_acc2'],
      spindleRunIn_backBearing_vibration_disp2 = _results[x]['spindleRunIn_backBearing_vibration_disp2'],
      )
      _objects.append(u)

    s.bulk_save_objects(_objects)

    try:
        s.commit()
    except pymysql.err.IntegrityError as e:
        s.rollback()
    except exc.IntegrityError as e:
        s.rollback()
    except Exception as e:
        s.rollback()
    #continue for loop
    print("spindle_runin and runin_data combine...")
    spindle_runin_record = s.query(SpindleRunIn).filter_by(id = spindle_runin_id).first()
    runin_data_records = s.query(RunInData).filter_by(spindleRunIn_id = spindle_runin_id).all()

    for array in runin_data_records:
      spindle_runin_record._runin_data.append(array)

    try:
        s.commit()
    except pymysql.err.IntegrityError as e:
        s.rollback()
    except exc.IntegrityError as e:
        s.rollback()
    except Exception as e:
        s.rollback()
    #end for loop

  s.close()

  if (return_message1 != '' or return_message2 != ''):
    return_message = return_message1 + '\n' + return_message2

  return jsonify({
    'status': return_value,
    'no_data': _no_spindle_or_user_data,
    'repeat_excel_files': _repeat_excel_files,
    'message': return_message,
  })


@excelTable.route("/listRunInFromCSV", methods=['POST'])
def list_runin_from_csv():
  print("listRunInFromCSV....")

  request_data = request.get_json()
  print("request_data: ", request_data)

  _file_name = request_data['file_name']
  #---
  #_base_dir = current_app.config['baseDir']

  load_dotenv()
  dotenv_path = "d:\\theta-asrs\\asrs\\.env\\.env"      #2024-04-26 add
  load_dotenv(dotenv_path)                              #2024-04-26 add
  _base_dir = os.getenv("baseDir")
  #---
  _path = _base_dir + '\\' + _file_name
  print("path: ", _path)

  _results = []
  _title_obj = {}
  return_value = True
  return_message = ''

  s = Session()
  existing_excel = s.query(SpindleRunIn).filter_by(spindleRunIn_excel_file=_file_name).first()
  s.close()

  workbook = openpyxl.load_workbook(filename = _path, read_only = True)

  if existing_excel or ('Sheet1' not in workbook.sheetnames):

    return_message = '錯誤! 已有相同Excel檔案名稱 或 沒有Sheet1...'
    return_value = False

    return jsonify({
      'title_obj': _title_obj,
      'outputs': _results,
      'status': return_value,
      'message': return_message,
    })

  #if 'Sheet1' in workbook.sheetnames:
  print('Sheet1 exists')
  #return_value = True
  sheet = workbook.active   #取得第1個工作表

  _customer_row = current_app.config['customer_row']
  _spindle_cat_row = current_app.config['spindle_cat_row']
  _id_row = current_app.config['id_row']
  _id_column = current_app.config['id_column']
  _emp_id_row = current_app.config['emp_id_row']
  _date_row = current_app.config['date_row']
  _date_column = current_app.config['date_column']
  _runin_row = current_app.config['runin_row']

  #spindleRunIn_excel_file = _file_name,      #excel file name
  spindleRunIn_customer = sheet.cell(row = _customer_row, column = 1).value       #客戶資料
  spindleRunIn_spindle_id = sheet.cell(row = _spindle_cat_row, column = 1).value     #主軸資料(spindle_cat)
  spindleRunIn_id = sheet.cell(row = _id_row, column = _id_column).value             #工單
  spindleRunIn_employee_id = str(sheet.cell(row = _emp_id_row, column = 1).value)   #員工資料
  spindleRunIn_date = str(sheet.cell(row = _date_row, column = _date_column).value.date())          #測試日期

  _title_obj = {
    'customer': spindleRunIn_customer,
    'spindle_cat': spindleRunIn_spindle_id,
    'run_in_id': spindleRunIn_id,
    'employer': spindleRunIn_employee_id,
    'date': spindleRunIn_date,
  }
  #print("spindleRunIn_excel_file: ", spindleRunIn_excel_file, _file_name)
  print("spindleRunIn_customer: ", spindleRunIn_customer)
  print("spindleRunIn_spindle_id: ", spindleRunIn_spindle_id)
  print("spindleRunIn_id: ", spindleRunIn_id)
  print("spindleRunIn_employee_id: ", spindleRunIn_employee_id)
  print("spindleRunIn_date: ", spindleRunIn_date)
  cnt=0
  for i in range(_runin_row, sheet.max_row+1):
    if sheet.cell(row = i, column = 1).value is None:
      #print("break i: ", i)
      break
    cnt +=1
    print("cnt, i: ", cnt, i)
    _obj = {
      'id': cnt,
      'spindleRunIn_period': sheet.cell(row = i, column = 1).value,         #時間
      'spindleRunIn_speed_level': sheet.cell(row = i, column = 2).value,    #段速
      'spindleRunIn_speed': sheet.cell(row = i, column = 3).value,          #轉速
      'spindleRunIn_stator_temp': sheet.cell(row = i, column = 4).value,    #定子溫度
      'spindleRunIn_inner_frontBearing_temp': sheet.cell(row = i, column = 5).value,   #內部前軸承溫度
      'spindleRunIn_inner_backBearing_temp': sheet.cell(row = i, column = 6).value,   #內部後軸承溫度
      'spindleRunIn_outer_frontBearing_temp': sheet.cell(row = i, column = 7).value,  #外部前軸承溫度
      'spindleRunIn_outer_backBearing_temp': sheet.cell(row = i, column = 8).value,   #外部後軸承溫度
      'spindleRunIn_room_temp': sheet.cell(row = i, column = 9).value,          #室溫
      'spindleRunIn_coolWater_temp': sheet.cell(row = i, column = 10).value,    #冷卻機水溫
      'spindleRunIn_Rphase_current': sheet.cell(row = i, column = 11).value,    #R相電流
      'spindleRunIn_Sphase_current': sheet.cell(row = i, column = 12).value,    #S相電流
      'spindleRunIn_Tphase_current': sheet.cell(row = i, column = 13).value,    #T相電流
      'spindleRunIn_cool_pipeline_flow': sheet.cell(row = i, column = 14).value,              #冷卻機管路流量
      'spindleRunIn_cool_pipeline_pressure': sheet.cell(row = i, column = 15).value,          #冷卻機管路壓力
      'spindleRunIn_frontBearing_vibration_speed1': sheet.cell(row = i, column = 16).value,   #前軸承震動速度-1
      'spindleRunIn_frontBearing_vibration_acc1': sheet.cell(row = i, column = 17).value,     #前軸承震動加速度-1
      'spindleRunIn_frontBearing_vibration_disp1': sheet.cell(row = i, column = 18).value,    #前軸承震動位移-1
      'spindleRunIn_frontBearing_vibration_speed2': sheet.cell(row = i, column = 19).value,   #前軸承震動速度-2
      'spindleRunIn_frontBearing_vibration_acc2': sheet.cell(row = i, column = 20).value,     #前軸承震動加速度-2
      'spindleRunIn_frontBearing_vibration_disp2': sheet.cell(row = i, column = 21).value,    #前軸承震動位移-2
      'spindleRunIn_backBearing_vibration_speed1': sheet.cell(row = i, column = 22).value,    #後軸承震動速度-1
      'spindleRunIn_backBearing_vibration_acc1': sheet.cell(row = i, column = 23).value,      #後軸承震動加速度-1
      'spindleRunIn_backBearing_vibration_disp1': sheet.cell(row = i, column = 24).value,     #後軸承震動位移-1
      'spindleRunIn_backBearing_vibration_speed2': sheet.cell(row = i, column = 25).value,    #後軸承震動速度-2
      'spindleRunIn_backBearing_vibration_acc2': sheet.cell(row = i, column = 26).value,      #後軸承震動加速度-2
      'spindleRunIn_backBearing_vibration_disp2': sheet.cell(row = i, column = 27).value,     #後軸承震動位移-2
    }

    _results.append(_obj)

    #print("_obj: ", return_value)

  return jsonify({
    'title_obj': _title_obj,
    'outputs': _results,
    'status': return_value,
    'message': return_message,
  })


@excelTable.route("/exportToCSVForStockInOut", methods=['POST'])
def export_to_csv_for_stock_in_out():
    print("exportToCSVForStockInOut....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    print("b data: ", _blocks, _count)

    return_value = True  # true: write into csv成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        for obj in _blocks:

          print("excel print: ", obj)

          if (obj['isIn']):
              obj['stockInTag_Employer'] = '入庫人員: ' + \
                  obj['stockInTag_Employer']
              obj['stockInTag_Date'] = '入庫日期: ' + \
                  obj['stockInTag_Date']
          else:
              obj['stockInTag_Employer'] = '領料人員: ' + \
                  obj['stockInTag_Employer']
              obj['stockInTag_Date'] = '領料日期: ' + \
                  obj['stockInTag_Date']

          obj['stockInTag_cnt'] = 1   # 2023-08-25 add

        print("a data: ", _blocks)

        for obj in _blocks:
            del obj['isIn']                 # 2022-12-23 remove 'isIn' key
            del obj['stockInTag_rePrint']   # 2023-08-10 remove 'stockInTag_rePrint' key
            del obj['stockInTag_reagName']  # 2023-08-10 remove 'stockInTag_reagName' key
        print("aa excel print: ", _blocks)

        # 查看當前工作目錄
        olddir = os.getcwd()
        print("當前工作目錄:%s" % olddir)

        myDrive = pathlib.Path.home().drive
        # print(myDrive)

        mypath = 'e:\\CMUHCH\\print.csv'
        #mypath = myDrive + '\\CMUHCH\\print.csv'
        myDir = 'e:\\CMUHCH\\'
        #myDir = myDrive + '\\CMUHCH\\'
        myFile = 'print.csv'

        if os.path.isdir(myDir):  #目錄存在
            if os.path.isfile(mypath):  #目錄內, 目標檔案存在
                os.chdir(myDir)         # 進入csv目錄
                # im.close()
                os.remove(myFile)       #刪除目標檔案
            else:
                #if not os.path.exists(myDir):
                #    os.makedirs(myDir)
                os.chdir(myDir)  # 進入csv目錄

            # 查看當前工作目錄
            curdir = os.getcwd()
            print("修改後工作目錄:%s" % curdir)

            # csvfile = open(myFile, 'a+')    # 使用 a+ 模式開啟檔案
            csvfile = open(myFile, 'w', newline='')
            try:
                fieldnames = ['name1', 'name2',
                              'stockInTag_reagID',
                              'stockInTag_batch',
                              'stockInTag_Date',
                              'stockInTag_Employer',
                              'stockInTag_reagTemp',
                              'stockInTag_alpha',
                              'stockInTag_cnt', ]  # 定義要寫入資料的鍵
                data = csv.DictWriter(
                    csvfile, fieldnames=fieldnames)  # 設定 data 為寫入資料
                data.writerows(_blocks)

            finally:
                csvfile.close()

            os.system("e:\\CMUHCH\\barcode.bat")

        else: #for test
          mypath = 'c:\\CMUHCH\\print.csv'
          myDir = 'c:\\CMUHCH\\'
          myFile = 'print.csv'

          os.chdir(myDir)         # 進入csv目錄
          if os.path.isfile(mypath):  #目錄內, 目標檔案存在
              os.remove(myFile)       #刪除目標檔案

          csvfile = open(myFile, 'w', newline='')
          try:
              fieldnames = ['name1', 'name2',
                            'stockInTag_reagID',
                            'stockInTag_batch',
                            'stockInTag_Date',
                            'stockInTag_Employer',
                            'stockInTag_reagTemp',
                            'stockInTag_alpha',
                            'stockInTag_cnt', ]  # 定義要寫入資料的鍵
              data = csv.DictWriter(
                  csvfile, fieldnames=fieldnames)  # 設定 data 為寫入資料
              data.writerows(_blocks)

          finally:
              csvfile.close()

        os.chdir(olddir)  # 進入server工作目錄
        print("最後工作目錄:%s" % olddir)

    return jsonify({
        'status': return_value,
        'outputs': mypath,
    })


@excelTable.route("/exportToExcelForReq", methods=['POST'])
def export_to_Excel_for_Req():
    print("exportToExcelForReq....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    #current_file = '..\\領用記錄查詢_'+today + '.xlsx'
    current_file = 'e:\\CMUHCH\\領用記錄查詢_'+today + '.xlsx'
    #current_file0 = '領用記錄查詢_'+today + '.xlsx'

    print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        if file_check:
            wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
        else:
            wb = Workbook()     # 建立空白的 Excel 活頁簿物件
        # ws = wb.active
        ws = wb.worksheets[0]   # 取得第一個工作表

        ws.title = '領用記錄查詢-' + _name                # 修改工作表 1 的名稱為 oxxo
        ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

        for obj in _blocks:
            temp_array = []
            temp_array.append(obj['reqRecord_reagID'])
            temp_array.append(obj['reqRecord_reagName'])
            temp_array.append(obj['reqRecord_supplier'])
            temp_array.append(obj['reqRecord_stockInDate'])
            temp_array.append(obj['reqRecord_Date'])
            temp_array.append(obj['reqRecord_Employer'])
            # temp_str = obj['reqRecord_cnt'] + obj['reqRecord_unit']
            temp_array.append(obj['reqRecord_cnt'])

            ws.append(temp_array)
            # print("obj: ", temp_array)

        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            temp_cell = column + '1'
            ws[temp_cell].font = Font(
                name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
            ws[temp_cell].alignment = Alignment(horizontal='center')
            ws.column_dimensions[column].bestFit = True
            # dim = openpyxl.worksheet.dimensions.ColumnDimension(
            #    ws, index=column, bestFit=True, customWidth=True)
            # ws.column_dimensions[column] = dim

        wb.save(current_file)

    myDrive = pathlib.Path.home().drive
    #mypath = 'e:\\CMUHCH\\print.csv'
    #mypath = myDrive + '\\CMUHCH\\print.csv'
    #myDir = 'e:\\CMUHCH\\'
    #myDir = myDrive + "\\CMUHCH\\" + current_file0
    #myDir = "e:\\CMUHCH\\" + current_file
    #print("myDir", myDir)

    return jsonify({
        'status': return_value,
        'outputs': current_file,
        # 'outputs': myDir,
    })


@excelTable.route("/exportToExcelForStorage", methods=['POST'])
def export_to_Excel_for_Storage():
    print("exportToExcelForStorage....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    #current_file = '..\\領用記錄查詢_'+today + '.xlsx'
    current_file = 'e:\\CMUHCH\\入庫記錄查詢_'+today + '.xlsx'
    #current_file0 = '領用記錄查詢_'+today + '.xlsx'

    print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        if file_check:
            wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
        else:
            wb = Workbook()     # 建立空白的 Excel 活頁簿物件
        # ws = wb.active
        ws = wb.worksheets[0]   # 取得第一個工作表

        ws.title = '入庫記錄查詢-' + _name                # 修改工作表 1 的名稱為 oxxo
        ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

        for obj in _blocks:
            temp_array = []
            temp_array.append(obj['reqRecord_reagID'])
            temp_array.append(obj['reqRecord_reagName'])
            temp_array.append(obj['reqRecord_supplier'])
            temp_array.append(obj['reqRecord_stockInDate'])
            #temp_array.append(obj['reqRecord_Date'])
            temp_array.append(obj['reqRecord_Employer'])
            # temp_str = obj['reqRecord_cnt'] + obj['reqRecord_unit']
            #temp_array.append(obj['reqRecord_cnt'])
            temp_array.append(obj['reqRecord_ori_count'])
            temp_array.append(obj['reqRecord_In_Unit'])

            ws.append(temp_array)
            # print("obj: ", temp_array)

        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            temp_cell = column + '1'
            ws[temp_cell].font = Font(
                name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
            ws[temp_cell].alignment = Alignment(horizontal='center')
            ws.column_dimensions[column].bestFit = True
            # dim = openpyxl.worksheet.dimensions.ColumnDimension(
            #    ws, index=column, bestFit=True, customWidth=True)
            # ws.column_dimensions[column] = dim

        wb.save(current_file)

    myDrive = pathlib.Path.home().drive
    #mypath = 'e:\\CMUHCH\\print.csv'
    #mypath = myDrive + '\\CMUHCH\\print.csv'
    #myDir = 'e:\\CMUHCH\\'
    #myDir = myDrive + "\\CMUHCH\\" + current_file0
    #myDir = "e:\\CMUHCH\\" + current_file
    #print("myDir", myDir)

    return jsonify({
        'status': return_value,
        'outputs': current_file,
        # 'outputs': myDir,
    })


@excelTable.route("/exportToExcelForStock", methods=['POST'])
def export_to_Excel_for_Stock():
    print("exportToExcelForStock....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']
    temp = len(_blocks)
    #data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    #current_file = '..\\庫存記錄查詢_'+today + '.xlsx'
    current_file = 'd:\\在庫記錄_' + today + '.xlsx'
    #current_file0 = '庫存記錄查詢_'+today + '.xlsx'

    print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功
    #if not data_check:  # false: 資料不完全
    #    return_value = False

    if return_value:
        if file_check:
            wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
        else:
            wb = Workbook()     # 建立空白的 Excel 活頁簿物件
        # ws = wb.active
        ws = wb.worksheets[0]   # 取得第一個工作表

        ws.title = '在庫記錄-' + _name                # 修改工作表 1 的名稱為 oxxo
        ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

        for obj in _blocks:
            temp_array = []

            temp_array.append(obj['spindleIn_workID'])
            temp_array.append(obj['spindleStockIn_type_cat'])
            temp_array.append(obj['spindleStockIn_date'])
            temp_array.append(obj['spindleStockIn_employer'])  # 入庫日期
            temp_array.append(obj['spindleStockIn_period'])  # 效期
            temp_array.append(obj['spindleStockOut_date'])
            temp_array.append(obj['spindleStockOut_employer'])
            temp_array.append(obj['spindleStockIn_st_lay'])
            temp_array.append(obj['comment'])
            temp_array.append(obj['date_comment'])

            ws.append(temp_array)
            # print("obj: ", temp_array)

        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            temp_cell = column + '1'
            ws[temp_cell].font = Font(
                name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
            ws[temp_cell].alignment = Alignment(horizontal='center')
            ws.column_dimensions[column].bestFit = True

        wb.save(current_file)

    myDrive = pathlib.Path.home().drive
    #mypath = 'e:\\CMUHCH\\print.csv'
    #mypath = myDrive + '\\CMUHCH\\print.csv'
    #myDir = 'e:\\CMUHCH\\'
    #myDir = myDrive + "\\CMUHCH\\" + current_file0
    #myDir = "e:\\CMUHCH\\" + current_file0
    #print("myDir", myDir)

    return jsonify({
        'status': return_value,
        'outputs': current_file,
        # 'outputs': myDir,
    })


@excelTable.route("/exportToExcelForInv", methods=['POST'])
def export_to_Excel_for_Inventory():
    print("exportToExcelForInv....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    #current_file = '..\\盤點作業_'+today + '.xlsx'
    current_file = 'e:\\盤點作業_'+today + '.xlsx'

    # print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        if file_check:
            wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
        else:
            wb = Workbook()     # 建立空白的 Excel 活頁簿物件
        # ws = wb.active
        ws = wb.worksheets[0]   # 取得第一個工作表

        ws.title = '庫存記錄查詢-' + _name        # 修改工作表 1 的名稱為 oxxo
        ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

        for obj in _blocks:
            print("exce obj: ", obj)
            temp_array = []
            temp_array.append(obj['id'])
            temp_array.append(obj['stockInTag_reagID'])
            temp_array.append(obj['stockInTag_reagName'])
            temp_array.append(obj['stockInTag_reagProduct'])  #類別, # 2023-05-23 add
            temp_array.append(obj['stockInTag_stockInBatch']) #批次, # 2023-02-16 add
            temp_array.append(obj['stockInTag_reagPeriod'])  # 效期
            temp_array.append(obj['stockInTag_reagTemp'])   # 保存溫度
            temp_array.append(obj['stockInTag_Date'])       # 入庫日期
            temp_array.append(obj['stockInTag_Employer'])   # 入庫人員

            temp_array.append(obj['stockInTag_grid'])       # 儲位
            temp_array.append(obj['stockInTag_cnt'])        # 在庫數
            temp_array.append(obj['stockInTag_cnt_inv_mdf'])  # 盤點數
            temp_array.append(obj['stockInTag_comment'])    # 說明

            ws.append(temp_array)

        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            temp_cell = column + '1'
            ws[temp_cell].font = Font(
                name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
            ws[temp_cell].alignment = Alignment(horizontal='center')
            ws.column_dimensions[column].bestFit = True

        wb.save(current_file)

    return jsonify({
        'status': return_value,
        'outputs': current_file,
    })
